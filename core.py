#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
core.py - crawling engine and link audit.

What it does:
  * collects EVERY page of a domain: sitemap.xml + robots.txt + live link crawl
    (not either/or, but both - that is how pages missing from the sitemap show up);
  * pulls links from <a> as well as <img>, <script>, <link>, <iframe>, <form>, ...
    (on many sites the buttons are not <a> tags at all);
  * searches for a given link/substring/regex across all pages, including raw HTML
    (catches links buried inside JS and JSON configs);
  * checks links for breakage and explains the REASON in plain language;
  * writes the result to .xlsx.

Used by app.py (web interface) and cli.py (command line).
"""

from __future__ import annotations

import html
import re
import threading
import time
from collections import deque
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from datetime import datetime
from typing import Callable, Iterable
from urllib.parse import urljoin, urlparse, urldefrag, urlunparse, parse_qsl, urlencode
import xml.etree.ElementTree as ET

import requests
from bs4 import BeautifulSoup

# ----------------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------------

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36 SiteLinkFinder/1.0"
)
TIMEOUT = 20
MAX_BODY_BYTES = 8 * 1024 * 1024  # never pull huge files into memory

# A timeout or a 429 usually means the site was busy, not that the page is
# broken. Retrying once keeps a hiccup from being reported as a dead link.
RETRIES = 1
RETRY_PAUSE = 1.5
RETRY_STATUSES = {429, 500, 502, 503, 504}

# Where links come from: tag -> attribute
LINK_SOURCES = {
    "a": "href",
    "area": "href",
    "link": "href",
    "img": "src",
    "script": "src",
    "iframe": "src",
    "frame": "src",
    "embed": "src",
    "source": "src",
    "video": "src",
    "audio": "src",
    "track": "src",
    "object": "data",
    "form": "action",
}

# What counts as a "real" link rather than a loaded resource
NAVIGATION_TAGS = {"a", "area", "form"}

# <link> rel values worth collecting at all
ALLOWED_LINK_RELS = {
    "stylesheet", "canonical", "alternate", "icon", "preload", "prefetch",
    "next", "prev", "amphtml", "shortlink",
}

# <link> rel values that point at PAGES rather than at resources. A broken
# hreflang or canonical is a real defect - search engines follow those tags -
# so they are checked alongside ordinary <a> links, not with images and css.
PAGE_LEVEL_RELS = {"canonical", "alternate", "next", "prev", "amphtml", "shortlink"}

# Pseudo-tag for links found in raw HTML (scripts, JSON, data attributes)
RAW_TAG = "page source"

SKIP_SCHEMES = ("mailto:", "tel:", "javascript:", "data:", "sms:", "callto:", "#")

# Extensions that should not be crawled as HTML pages
NON_PAGE_EXT = re.compile(
    r"\.(jpg|jpeg|png|gif|svg|webp|ico|bmp|avif|css|js|mjs|json|xml|txt|pdf|zip|rar|7z|"
    r"gz|tar|mp4|webm|mp3|wav|avi|mov|woff2?|ttf|eot|otf|doc|docx|xls|xlsx|ppt|pptx|csv)$",
    re.I,
)

# Links hidden in raw HTML / JS.
# Absolute: https://...  Relative: "/path" inside quotes.
RAW_URL_RE = re.compile(r"""https?://[^\s"'<>\\)\]]+""", re.I)
RAW_REL_RE = re.compile(r"""["'(](/[A-Za-z0-9\-._~!$&*+,;=:@%/?#\[\]]{2,300})["')]""")

STATUS_REASONS = {
    400: "Bad request (400)",
    401: "Authorization required (401)",
    403: "Forbidden (403) - often bot protection, check manually",
    404: "Page not found (404) - broken link",
    405: "Method not allowed (405)",
    408: "Request timeout (408)",
    410: "Page permanently removed (410)",
    429: "Too many requests (429) - lower the thread count",
    451: "Blocked for legal reasons (451)",
    500: "Internal server error (500)",
    502: "Bad gateway (502)",
    503: "Service unavailable (503)",
    504: "Gateway timeout (504)",
    # LinkedIn answers any bot with 999; the link itself is almost always fine
    999: "Bot protection (999) - LinkedIn blocks crawlers, the link is likely fine",
}


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------

def normalize_domain(domain: str) -> str:
    """'example.com' -> 'https://example.com'"""
    domain = (domain or "").strip()
    if not domain:
        return ""
    if not domain.startswith(("http://", "https://")):
        domain = "https://" + domain
    return domain.rstrip("/")


def registrable(netloc: str) -> str:
    """Rough 'example.com' out of 'www.blog.example.com' - used for subdomain checks."""
    host = netloc.lower().split(":")[0]
    parts = host.split(".")
    return ".".join(parts[-2:]) if len(parts) >= 2 else host


def clean_url(url: str) -> str:
    """Strips the fragment and stray whitespace/newlines out of an href."""
    url = html.unescape((url or "").strip())
    url = re.sub(r"\s+", "", url)
    url, _ = urldefrag(url)
    return url


# Advertising and analytics tags: they do not change page content but spawn
# "new" addresses during a crawl (?_ga=..., ?utm_source=... and friends)
# NB: only unambiguous tracking keys belong here. A generic name like "ref" is
# skipped on purpose - on some sites (?ref=main) it selects real content, and
# collapsing it would merge pages that are genuinely different.
TRACKING_PARAMS = {
    "_ga", "_gl", "_hsenc", "_hsmi", "hsctatracking", "gclid", "dclid", "fbclid",
    "msclkid", "yclid", "igshid", "mc_cid", "mc_eid", "referrer",
    "vero_id", "vero_conv", "s_kwcid", "twclid", "ttclid", "li_fat_id",
}


def strip_tracking(url: str) -> str:
    """Removes tracking tags from an address, keeping meaningful parameters."""
    try:
        p = urlparse(url)
    except Exception:
        return url
    if not p.query:
        return url
    kept = [
        (k, v) for k, v in parse_qsl(p.query, keep_blank_values=True)
        if k.lower() not in TRACKING_PARAMS and not k.lower().startswith("utm_")
    ]
    return urlunparse(p._replace(query=urlencode(kept)))


def url_key(url: str) -> str:
    """
    Normalized key for comparing links: scheme, www, trailing slash, host case
    and tracking tags are all ignored.
    """
    try:
        p = urlparse(strip_tracking(url))
    except Exception:
        return (url or "").lower()
    host = p.netloc.lower().split(":")[0]
    if host.startswith("www."):
        host = host[4:]
    path = p.path or "/"
    if len(path) > 1:
        path = path.rstrip("/")
    # sort parameters: ?a=1&b=2 and ?b=2&a=1 are the same address
    pairs = sorted(parse_qsl(p.query, keep_blank_values=True))
    q = f"?{urlencode(pairs)}" if pairs else ""
    return f"{host}{path}{q}".lower()


HOSTLIKE_RE = re.compile(r"^[a-z0-9](?:[a-z0-9-]*[a-z0-9])?(?:\.[a-z0-9-]+)+(?:[:/?#]|$)", re.I)


def query_to_url(query: str, root: str) -> str:
    """
    Turns whatever the user typed into an absolute URL.

    'https://site.com/page'  -> unchanged
    '/page'                  -> https://site.com/page
    'site.com/page'          -> https://site.com/page   (not site.com/site.com/page)
    'page'                   -> https://site.com/page
    """
    q = clean_url(query)
    if not q:
        return ""
    if q.lower().startswith(("http://", "https://")):
        return q
    if q.startswith("//"):
        return "https:" + q
    if HOSTLIKE_RE.match(q):
        return "https://" + q
    return urljoin(root.rstrip("/") + "/", q.lstrip("/"))


def pl(n: int, word: str, plural_form: str | None = None) -> str:
    """1 page / 2 pages"""
    return word if abs(n) == 1 else (plural_form or word + "s")


class RobotsRules:
    """
    robots.txt parser.

    Python's urllib.robotparser is not usable here: a blank line after
    "User-agent: *" ends the record for it, so a perfectly ordinary file like

        User-agent: *

        # comment
        Disallow: /private

    parses into zero rules and every URL comes back allowed. Real files are
    written that way all the time, so the check silently passed everything.

    This follows RFC 9309 instead: blank lines and comments are ignored, a
    group ends only at the next User-agent line, and Allow/Disallow support
    the '*' and '$' wildcards. The longest matching rule wins, with Allow
    beating Disallow on a tie - the behaviour Google documents.
    """

    def __init__(self, text: str = ""):
        self.rules: list[tuple[str, bool]] = []   # (pattern, allowed)
        self.disallow_count = 0
        self._cache: dict[str, bool] = {}
        if text:
            self.parse(text)

    def parse(self, text: str) -> None:
        applies = False          # are we inside a group that covers everyone?
        expecting_agents = False  # consecutive User-agent lines share one group

        for raw in text.splitlines():
            line = raw.split("#", 1)[0].strip()
            if not line or ":" not in line:
                continue
            field, _, value = line.partition(":")
            field, value = field.strip().lower(), value.strip()

            if field == "user-agent":
                if not expecting_agents:      # a new group starts here
                    applies = False
                    expecting_agents = True
                if value == "*":
                    applies = True
                continue

            expecting_agents = False
            if not applies or field not in ("allow", "disallow"):
                continue
            if field == "disallow" and not value:
                continue                      # "Disallow:" alone means allow all
            self.rules.append((value, field == "allow"))
            if field == "disallow":
                self.disallow_count += 1

    @staticmethod
    def _matches(pattern: str, path: str) -> bool:
        anchored = pattern.endswith("$")
        if anchored:
            pattern = pattern[:-1]
        regex = "".join(".*" if ch == "*" else re.escape(ch) for ch in pattern)
        return bool(re.match(regex + ("$" if anchored else ""), path))

    def allowed(self, url: str) -> bool:
        """May a crawler fetch this URL?"""
        if not self.rules:
            return True
        try:
            p = urlparse(url)
        except Exception:
            return True
        path = (p.path or "/") + (f"?{p.query}" if p.query else "")

        cached = self._cache.get(path)
        if cached is not None:
            return cached

        best_len, verdict = -1, True
        for pattern, is_allow in self.rules:
            if self._matches(pattern, path) and len(pattern) >= best_len:
                # equal length: Allow wins, as Google specifies
                if len(pattern) > best_len or is_allow:
                    best_len, verdict = len(pattern), is_allow
        self._cache[path] = verdict
        return verdict


def is_page_like(url: str) -> bool:
    """Does this look like an HTML page (rather than an image or archive)?"""
    path = urlparse(url).path
    return not NON_PAGE_EXT.search(path)


def describe_status(status) -> str:
    """Human-readable reason why a link is considered broken."""
    if isinstance(status, str):
        return status
    if status in STATUS_REASONS:
        return STATUS_REASONS[status]
    if 300 <= status < 400:
        return f"Redirect ({status})"
    if 400 <= status < 500:
        return f"Client error ({status})"
    if status >= 500:
        return f"Server error ({status})"
    return f"OK ({status})"


def describe_exception(exc: Exception) -> str:
    """Network errors explained in plain language."""
    import requests.exceptions as rex

    if isinstance(exc, rex.SSLError):
        return "SSL certificate error (site has invalid HTTPS)"
    if isinstance(exc, rex.ConnectTimeout):
        return "Connection timeout - server is not responding"
    if isinstance(exc, rex.ReadTimeout):
        return "Read timeout - server started replying but never finished"
    if isinstance(exc, rex.TooManyRedirects):
        return "Redirect loop (the link keeps bouncing in circles)"
    if isinstance(exc, rex.InvalidURL) or isinstance(exc, rex.MissingSchema):
        return "Malformed URL (typo in the href)"
    if isinstance(exc, rex.ConnectionError):
        text = str(exc)
        if "NameResolutionError" in text or "getaddrinfo" in text or "Name or service" in text:
            return "Domain does not exist / DNS does not resolve"
        if "RemoteDisconnected" in text or "ConnectionResetError" in text:
            return "Server dropped the connection"
        return "Could not connect to the server"
    return f"Error: {type(exc).__name__}: {str(exc)[:150]}"


# ----------------------------------------------------------------------------
# Data model
# ----------------------------------------------------------------------------

# How to explain where exactly a link was found and whether it is visible
SOURCE_LABELS = {
    "a": ("link on the page", True),
    "area": ("clickable area on an image map", True),
    "form": ("form (submit target)", True),
    "img": ("image", True),
    "iframe": ("embedded iframe", True),
    "frame": ("frame", True),
    "video": ("video", True),
    "audio": ("audio", True),
    "source": ("media source", True),
    "track": ("subtitles", False),
    "embed": ("embedded object", True),
    "object": ("embedded object", True),
    "script": ("script (not visible on the page)", False),
}

LINK_REL_LABELS = {
    "canonical": ("canonical in <head> - technical, not visible on the page", False),
    "alternate": ("hreflang in <head> - SEO tag, NOT visible on the page", False),
    "stylesheet": ("CSS stylesheet (not visible on the page)", False),
    "icon": ("site icon", False),
    "preload": ("resource preload", False),
    "prefetch": ("resource prefetch", False),
    "dns-prefetch": ("browser hint", False),
    "preconnect": ("browser hint", False),
    "amphtml": ("AMP version of the page", False),
    "next": ("pagination link", False),
    "prev": ("pagination link", False),
}


MATCH_KINDS = {
    "direct": "links here",
    "mention": "only mentions the address inside another link",
    "self": "the page referring to itself",
}


def points_at_page(tag: str, rel: str = "") -> bool:
    """
    Does this link point at a page (rather than at a loaded resource)?

    True for <a>/<area>/<form> and for <head> tags that reference pages:
    hreflang, canonical, pagination. False for stylesheets, icons, images.
    """
    if tag in NAVIGATION_TAGS:
        return True
    if tag == "link":
        return any(r in PAGE_LEVEL_RELS for r in (rel or "").lower().split())
    return False


def describe_source(tag: str, rel: str = "") -> tuple[str, bool]:
    """
    Plain-language explanation of where a link was found.
    Returns (description, whether it is visible on the page).
    """
    if tag == RAW_TAG:
        return ("inside page source (script/JSON) - not visible on the page", False)
    if tag == "link":
        for key in (rel or "").lower().split():
            if key in LINK_REL_LABELS:
                return LINK_REL_LABELS[key]
        return (f"<link{' rel=' + rel if rel else ''}> tag in <head> - technical", False)
    return SOURCE_LABELS.get(tag, (f"<{tag}> tag", True))


@dataclass
class LinkHit:
    page: str          # page it was found on
    href: str          # exactly as written in the code
    absolute: str      # absolute URL
    text: str          # link text / alt / caption
    tag: str           # a, img, link, script, RAW_TAG ...
    context: str = ""  # surrounding snippet (for source matches)
    rel: str = ""      # rel attribute of <link> - it defines the meaning
    status: object = None  # status of the found link itself (filled during search)
    no_internal: bool = False  # nothing but technical tags points at this address

    # How the match relates to the address we searched for:
    #   direct  - the link actually points there
    #   mention - the address only appears inside another link, e.g. a share
    #             button: facebook.com/share.php?u=<our address>
    #   self    - the page links to itself (canonical, hreflang, share button)
    kind: str = "direct"

    # Whether the page holding the link is canonical. Links from a page whose
    # canonical points elsewhere (typically pagination) carry no weight for
    # search engines, and crawlers like Sitebulb leave them out of the count.
    source_canonical: bool = True

    # Indexability of the two pages involved, when we crawled them. Empty for
    # addresses outside the domain - we never fetched those, so we do not guess.
    target_index_status: str = ""   # can the address the link points at be indexed?
    target_index_reason: str = ""
    source_index_status: str = ""   # can the page holding the link be indexed?
    source_index_reason: str = ""

    @property
    def where(self) -> str:
        return describe_source(self.tag, self.rel)[0]

    @property
    def visible(self) -> bool:
        return describe_source(self.tag, self.rel)[1]

    @property
    def nofollow(self) -> bool:
        """The site tells engines not to pass any weight through this link."""
        parts = (self.rel or "").lower().split()
        return any(k in parts for k in ("nofollow", "sponsored", "ugc"))


@dataclass
class PageInfo:
    url: str
    status: object = None
    final_url: str = ""
    title: str = ""
    content_type: str = ""
    links_count: int = 0
    error: str = ""
    canonical: str = ""      # rel=canonical, if the page declares one
    meta_robots: str = ""    # <meta name="robots" content="...">
    x_robots: str = ""       # X-Robots-Tag response header
    indexable: bool = True   # may a search engine index this page?
    index_status: str = "Indexable"   # short verdict for the table
    index_reason: str = ""            # what exactly decided it
    inbound: int = 0         # how many pages of this site link TO this page


@dataclass
class Options:
    domain: str
    mode: str = "search"                  # search | broken | pages | full
    query: str = ""                       # what to look for
    match: str = "contains"               # contains | exact | regex
    limit: int | None = None              # max pages
    max_depth: int | None = None          # crawl depth
    workers: int = 10
    delay: float = 0.0                    # pause between requests, seconds
    include_subdomains: bool = False
    use_sitemap: bool = True
    use_crawl: bool = True
    check_external: bool = False          # check external links for breakage
    check_assets: bool = False            # check images/scripts/css
    find_orphans: bool = False            # also list pages nothing links to
    only_non_indexable: bool = False      # show just the pages engines will skip
    respect_robots: bool = False          # skip URLs robots.txt disallows
    search_raw_html: bool = True          # also search raw HTML (JS, JSON)
    exclude: list[str] = field(default_factory=list)  # URL patterns to skip


# ----------------------------------------------------------------------------
# Main class
# ----------------------------------------------------------------------------

class SiteAuditor:
    def __init__(
        self,
        opts: Options,
        on_log: Callable[[str], None] | None = None,
        on_progress: Callable[[str, int, int], None] | None = None,
        stop_event: threading.Event | None = None,
    ):
        self.opts = opts
        self.root = normalize_domain(opts.domain)
        self.root_netloc = urlparse(self.root).netloc
        self.root_reg = registrable(self.root_netloc)

        self._log_cb = on_log or (lambda m: None)
        self._progress_cb = on_progress or (lambda p, d, t: None)
        self.stop_event = stop_event or threading.Event()

        self._local = threading.local()
        self._lock = threading.Lock()

        # results
        self.pages: dict[str, PageInfo] = {}
        self.all_links: list[LinkHit] = []
        self.search_hits: list[LinkHit] = []
        self.broken: list[dict] = []
        self.orphans: list[dict] = []
        self.status_cache: dict[str, object] = {}
        self.sitemap_count = 0
        self.sitemap_keys: set[str] = set()
        self.inbound_pages = 0            # pages that really link to the query
        self.inbound_pages_canonical = 0  # ... of which are canonical
        self._robots: RobotsRules | None = None
        self._robots_text: str | None = None   # fetched once, reused
        self._robots_skipped: set[str] = set()
        self.started_at = None
        self.finished_at = None

    # -- infrastructure ------------------------------------------------------

    def log(self, msg: str) -> None:
        self._log_cb(msg)

    def progress(self, phase: str, done: int, total: int) -> None:
        self._progress_cb(phase, done, total)

    @property
    def session(self) -> requests.Session:
        """One session per thread - keep-alive speeds the crawl up massively."""
        s = getattr(self._local, "session", None)
        if s is None:
            s = requests.Session()
            s.headers.update({
                "User-Agent": USER_AGENT,
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
            })
            self._local.session = s
        return s

    def stopped(self) -> bool:
        return self.stop_event.is_set()

    # -- filters -------------------------------------------------------------

    def in_scope(self, url: str) -> bool:
        """Does this URL belong to our domain?"""
        try:
            netloc = urlparse(url).netloc.lower()
        except Exception:
            return False
        if not netloc:
            return False
        if self.opts.include_subdomains:
            return registrable(netloc) == self.root_reg
        a, b = netloc.split(":")[0], self.root_netloc.lower().split(":")[0]
        return a == b or a == "www." + b or "www." + a == b

    def excluded(self, url: str) -> bool:
        if any(pat and pat.lower() in url.lower() for pat in self.opts.exclude):
            return True
        if self.opts.respect_robots:
            self.load_robots()
            if self._robots and not self._robots.allowed(url):
                # the same address is checked once per link pointing at it,
                # so count distinct addresses rather than checks
                self._robots_skipped.add(url_key(url))
                return True
        return False

    # -- source discovery ----------------------------------------------------

    def sitemap_candidates(self) -> list[str]:
        """Sitemaps from robots.txt plus the standard paths."""
        found = []
        robots = urljoin(self.root + "/", "robots.txt")
        try:
            r = self.session.get(robots, timeout=TIMEOUT)
            # keep the text: the indexability check needs the same file, and
            # fetching it twice would be wasteful
            self._robots_text = r.text if r.status_code == 200 else ""
            if r.status_code == 200:
                for line in r.text.splitlines():
                    if line.lower().startswith("sitemap:"):
                        sm = line.split(":", 1)[1].strip()
                        if sm:
                            found.append(sm)
                if found:
                    self.log(f"  robots.txt points to {len(found)} sitemap "
                             f"{pl(len(found), 'file')}")
        except Exception:
            pass
        for name in ("sitemap.xml", "sitemap_index.xml", "sitemap-index.xml", "sitemap/sitemap.xml"):
            found.append(urljoin(self.root + "/", name))
        # de-duplicate, keep order
        seen, out = set(), []
        for u in found:
            if u not in seen:
                seen.add(u)
                out.append(u)
        return out

    def collect_sitemap_urls(self) -> list[str]:
        """Recursively expands sitemap indexes (common on larger sites)."""
        found: set[str] = set()
        to_check = deque(self.sitemap_candidates())
        checked: set[str] = set()

        while to_check and not self.stopped():
            sm_url = to_check.popleft()
            if sm_url in checked or len(checked) > 200:
                continue
            checked.add(sm_url)
            try:
                r = self.session.get(sm_url, timeout=TIMEOUT)
                if r.status_code != 200:
                    continue
                root_el = ET.fromstring(r.content)
            except ET.ParseError:
                continue
            except Exception:
                continue

            tag = root_el.tag.split("}")[-1]
            locs = [
                el.text.strip()
                for el in root_el.iter()
                if el.tag.split("}")[-1] == "loc" and el.text
            ]
            if tag == "sitemapindex":
                self.log(f"  sitemap index: {sm_url} -> {len(locs)} nested")
                for loc in locs:
                    if loc not in checked:
                        to_check.append(loc)
            else:
                before = len(found)
                found.update(locs)
                if len(found) > before:
                    self.log(f"  {sm_url}: +{len(found) - before} URLs")

        return sorted(found)

    # -- page download -------------------------------------------------------

    def fetch(self, url: str, attempt: int = 0) -> tuple[PageInfo, str]:
        """Downloads a page. Returns (PageInfo, html text)."""
        info = PageInfo(url=url)
        try:
            r = self.session.get(url, timeout=TIMEOUT, allow_redirects=True, stream=True)
            if r.status_code in RETRY_STATUSES and attempt < RETRIES and not self.stopped():
                r.close()
                time.sleep(RETRY_PAUSE)
                return self.fetch(url, attempt + 1)
            info.status = r.status_code
            info.final_url = r.url
            ctype = r.headers.get("Content-Type", "")
            info.content_type = ctype.split(";")[0].strip()
            info.x_robots = r.headers.get("X-Robots-Tag", "")

            if "html" not in ctype.lower():
                r.close()
                return info, ""

            chunks, size = [], 0
            for chunk in r.iter_content(65536):
                chunks.append(chunk)
                size += len(chunk)
                if size > MAX_BODY_BYTES:
                    break
            r.close()
            encoding = r.encoding or r.apparent_encoding or "utf-8"
            text = b"".join(chunks).decode(encoding, errors="replace")
            return info, text
        except Exception as e:
            import requests.exceptions as rex
            transient = isinstance(e, (rex.Timeout, rex.ConnectionError))
            if transient and attempt < RETRIES and not self.stopped():
                time.sleep(RETRY_PAUSE)
                return self.fetch(url, attempt + 1)
            info.status = "ERROR"
            info.error = describe_exception(e)
            return info, ""

    def extract_links(self, page_url: str, html_text: str) -> tuple[list[LinkHit], str, str]:
        """Pulls every link off a page. Returns (links, title, meta robots)."""
        hits: list[LinkHit] = []
        try:
            soup = BeautifulSoup(html_text, "html.parser")
        except Exception:
            return hits, "", ""

        title = soup.title.get_text(strip=True)[:200] if soup.title else ""

        # <meta name="robots"> and the Google-specific variant decide indexing
        meta_robots = " ".join(
            (m.get("content") or "").strip()
            for m in soup.find_all("meta")
            if (m.get("name") or "").lower() in ("robots", "googlebot")
        ).strip()

        # <base href> re-points every relative link on the page. Miss it and
        # each one resolves against the wrong folder. It only affects how links
        # are resolved - the page keeps its own address.
        base_url = page_url
        base_tag = soup.find("base", href=True)
        if base_tag:
            candidate = urljoin(page_url, clean_url(base_tag["href"]))
            if candidate.lower().startswith(("http://", "https://")):
                base_url = candidate

        for tag_name, attr in LINK_SOURCES.items():
            for el in soup.find_all(tag_name):
                raw = el.get(attr)
                if not raw or not isinstance(raw, str):
                    continue
                raw = raw.strip()
                if not raw or raw.lower().startswith(SKIP_SCHEMES):
                    continue
                # rel matters on every tag: on <link> it says what the tag is
                # for, on <a> it can say nofollow/sponsored/ugc - links the site
                # deliberately refuses to vouch for
                rel = " ".join(el.get("rel") or []).lower()
                if tag_name == "link" and not any(k in ALLOWED_LINK_RELS for k in rel.split()):
                    continue
                cleaned = clean_url(raw)
                if not cleaned:
                    continue
                try:
                    absolute = urljoin(base_url, cleaned)
                except Exception:
                    continue
                if not absolute.lower().startswith(("http://", "https://")):
                    continue

                text = (
                    el.get_text(strip=True)[:200]
                    if tag_name in ("a", "area")
                    else (el.get("alt") or el.get("title") or "")[:200]
                )
                # for hreflang tags it helps to see which language it points at
                if tag_name == "link" and el.get("hreflang"):
                    text = f"hreflang={el.get('hreflang')}"

                hits.append(LinkHit(
                    page=page_url, href=raw, absolute=absolute,
                    text=text, tag=tag_name, rel=rel,
                ))
        return hits, title, meta_robots

    # -- phase 1: crawl ------------------------------------------------------

    def crawl(self) -> None:
        seeds: list[str] = [self.root]

        if self.opts.use_sitemap:
            self.log("Looking for sitemap.xml and robots.txt ...")
            sm = self.collect_sitemap_urls()
            in_scope = [u for u in sm if self.in_scope(u)]
            self.log(f"Sitemap gave {len(in_scope)} pages on this domain "
                     f"({len(sm)} entries in total)")
            self.sitemap_count = len(in_scope)
            self.sitemap_keys = {url_key(u) for u in in_scope}
            seeds.extend(in_scope)
        if not self.opts.use_crawl and len(seeds) == 1:
            self.log("Sitemap is empty and link crawling is off - enabling the crawl anyway.")
            self.opts.use_crawl = True

        # queue: (url, depth)
        queue: deque[tuple[str, int]] = deque()
        queued: set[str] = set()
        for s in seeds:
            k = url_key(s)
            if k not in queued and not self.excluded(s):
                queued.add(k)
                queue.append((clean_url(s), 0))

        self.log(f"Starting queue: {len(queue)} addresses. Crawling ...")
        limit = self.opts.limit
        processed = 0
        discovered = 0     # new addresses queued since the last report
        explained = False  # the note about the growing queue is printed once

        def handle(item: tuple[str, int]) -> tuple[PageInfo, list[LinkHit], int]:
            url, depth = item
            if self.opts.delay:
                time.sleep(self.opts.delay)
            info, text = self.fetch(url)
            links, title, meta_robots = (
                self.extract_links(url, text) if text else ([], "", "")
            )
            info.title = title
            info.meta_robots = meta_robots
            info.links_count = len(links)
            info.canonical = next(
                (l.absolute for l in links
                 if l.tag == "link" and "canonical" in (l.rel or "")),
                "",
            )
            if self.opts.search_raw_html and text and self.opts.query:
                known = {url_key(l.absolute) for l in links}
                links.extend(self.scan_raw_html(url, text, known))
            return info, links, depth

        with ThreadPoolExecutor(max_workers=self.opts.workers) as pool:
            while queue and not self.stopped():
                if limit and processed >= limit:
                    self.log(f"Reached the limit of {limit} pages - stopping the crawl.")
                    break

                batch: list[tuple[str, int]] = []
                batch_size = self.opts.workers * 3
                while queue and len(batch) < batch_size:
                    if limit and processed + len(batch) >= limit:
                        break
                    batch.append(queue.popleft())
                if not batch:
                    break

                for info, links, depth in pool.map(handle, batch):
                    processed += 1
                    self.pages[info.url] = info
                    self.status_cache[info.url] = info.status if not info.error else info.error

                    with self._lock:
                        self.all_links.extend(links)

                    # only navigation links extend the queue
                    if self.opts.use_crawl and not self.stopped():
                        if self.opts.max_depth is not None and depth >= self.opts.max_depth:
                            continue
                        for lk in links:
                            if lk.tag not in NAVIGATION_TAGS:
                                continue
                            tgt = lk.absolute
                            if not self.in_scope(tgt) or self.excluded(tgt):
                                continue
                            if not is_page_like(tgt):
                                continue
                            k = url_key(tgt)
                            if k in queued:
                                continue
                            queued.add(k)
                            # queue the address without tracking tags so the same
                            # page is never downloaded twice
                            queue.append((strip_tracking(tgt), depth + 1))
                            discovered += 1

                total_known = processed + len(queue)
                self.progress("crawl", processed, min(total_known, limit) if limit else total_known)

                if discovered and not explained:
                    explained = True
                    self.log("  (crawled pages link to pages that were not on the list yet - "
                             "those get added to the queue, which is why it grows at first)")
                added = (
                    f", queued {discovered} more {pl(discovered, 'address', 'addresses')}"
                    if discovered else ""
                )
                self.log(f"  pages crawled: {processed}, left in queue: {len(queue)}{added}")
                discovered = 0

        self.load_robots()   # cheap: the file was already fetched for the sitemap
        if self.opts.respect_robots and self._robots_skipped:
            n = len(self._robots_skipped)
            self.log(f"Skipped {n} {pl(n, 'address', 'addresses')} because "
                     f"robots.txt disallows them")
        elif self._robots and self._robots.disallow_count and not self.opts.respect_robots:
            self.log(f"Note: robots.txt has {self._robots.disallow_count} Disallow "
                     f"{pl(self._robots.disallow_count, 'rule')}; crawling those "
                     f"addresses anyway (tick 'Respect robots.txt' to skip them)")

        extra = len(queued) - self.sitemap_count
        if self.opts.use_crawl and extra > 0:
            self.log(f"Beyond the sitemap, the crawl found {extra} more {pl(extra, 'page')}.")
        self.log(f"Crawl finished: {len(self.pages)} {pl(len(self.pages), 'page')}, "
                 f"{len(self.all_links)} links collected.")

    # -- phase 2: search -----------------------------------------------------

    def scan_raw_html(self, page_url: str, text: str, known: set[str]) -> list[LinkHit]:
        """
        Searches the HTML source directly - catches links inside <script>, JSON
        and data attributes where there is no <a> tag at all.

        We do not match "a substring anywhere", only URL-like chunks: otherwise
        a query such as 'privacy' would match every word in the page text.
        """
        q = self.opts.query
        if not q:
            return []

        try:
            rx = re.compile(q, re.I) if self.opts.match == "regex" else None
        except re.error:
            return []
        target_key = url_key(query_to_url(q, self.root))
        needle = q.lower()

        def matches(candidate: str) -> bool:
            if rx is not None:
                return bool(rx.search(candidate))
            if self.opts.match == "exact":
                return url_key(urljoin(page_url, candidate)) == target_key
            return needle in candidate.lower()

        hits: list[LinkHit] = []
        seen: set[str] = set()

        for m in list(RAW_URL_RE.finditer(text)) + list(RAW_REL_RE.finditer(text)):
            raw = (m.group(1) if m.lastindex else m.group(0)).rstrip(".,;)")
            raw = html.unescape(raw)
            if not matches(raw):
                continue
            try:
                absolute = urljoin(page_url, clean_url(raw))
            except Exception:
                continue
            key = url_key(absolute)
            # if a tag on this same page already gave us this link, skip it
            if key in known or key in seen:
                continue
            seen.add(key)
            start, end = max(0, m.start() - 80), min(len(text), m.end() + 80)
            context = re.sub(r"\s+", " ", text[start:end]).strip()
            hits.append(LinkHit(
                page=page_url, href=raw, absolute=absolute,
                text="", tag=RAW_TAG, context=context,
            ))
            if len(hits) >= 15:
                break
        return hits

    def classify_hits(self, needle: str) -> None:
        """
        Sorts matches into real links, mere mentions and self-references, and
        marks which source pages are canonical.

        A share button reads facebook.com/share.php?u=<our address>: the address
        is in the query string of a link that goes to Facebook. It matches the
        search, but it is not a link to us - counting it as one is how inflated
        "internal links" numbers happen.
        """
        for h in self.search_hits:
            target = urlparse(h.absolute)
            in_address = needle in (target.netloc + target.path).lower()
            query = (target.query or "").lower()
            # A share button carries a whole URL inside its query string. A
            # match in the query only means "not a link here" when the query
            # actually holds another address - otherwise the query is just this
            # link's own parameter, e.g. ?hs_preview=..., and the link is real.
            carries_url = bool(re.search(r"https?(://|%3a%2f%2f)", query))
            in_query = bool(query) and needle in query

            if in_query and not in_address and carries_url:
                h.kind = "mention"
            elif url_key(h.page) == url_key(h.absolute):
                h.kind = "self"
            else:
                h.kind = "direct"

            src = self.pages.get(h.page)
            if src and src.canonical:
                h.source_canonical = url_key(src.canonical) == url_key(h.page)

        # Attach the indexability verdicts we already worked out. "Works" and
        # "will show up in search" are different things: a link can lead to a
        # perfectly healthy page that no engine will ever index.
        by_key = {url_key(url): page for url, page in self.pages.items()}
        for h in self.search_hits:
            source = by_key.get(url_key(h.page))
            if source:
                h.source_index_status = source.index_status
                h.source_index_reason = source.index_reason
            target = by_key.get(url_key(h.absolute))
            if target:
                h.target_index_status = target.index_status
                h.target_index_reason = target.index_reason

    def run_search(self) -> None:
        q = self.opts.query
        if not q:
            return
        mode = self.opts.match
        self.log(f"Searching for '{q}' (mode: {mode}) across {len(self.all_links)} links ...")

        if mode == "regex":
            try:
                rx = re.compile(q, re.I)
            except re.error as e:
                self.log(f"Invalid regular expression: {e}")
                return
        target_key = url_key(query_to_url(q, self.root))
        needle = q.lower()
        if mode == "exact":
            self.log(f"  exact match resolved to: {query_to_url(q, self.root)}")

        seen: set[tuple] = set()
        for lk in self.all_links:
            if lk.tag == RAW_TAG:
                matched = True  # already filtered during the source scan
            elif mode == "exact":
                matched = url_key(lk.absolute) == target_key
            elif mode == "regex":
                matched = bool(rx.search(lk.href) or rx.search(lk.absolute) or rx.search(lk.text or ""))
            else:
                matched = needle in lk.href.lower() or needle in lk.absolute.lower()

            if matched:
                sig = (lk.page, lk.href, lk.absolute, lk.tag, lk.context[:60])
                if sig in seen:
                    continue
                seen.add(sig)
                self.search_hits.append(lk)

        pages_with = len({h.page for h in self.search_hits})
        self.log(f"Found {len(self.search_hits)} {pl(len(self.search_hits), 'match', 'matches')} "
                 f"on {pages_with} {pl(pages_with, 'page')}.")

        if not self.search_hits:
            return

        self.classify_hits(needle)

        # Per address: is it referenced by any normal link at all? If every
        # match sits in a technical tag, nothing on the site actually links
        # there - the SEO tools call this "no internal linking URLs".
        # Only a real link from another page counts. A share button that carries
        # the address in its query string does not, nor does the page itself.
        visible_per_target: dict[str, int] = {}
        for h in self.search_hits:
            key = url_key(h.absolute)
            counts = h.visible and h.kind == "direct"
            visible_per_target[key] = visible_per_target.get(key, 0) + (1 if counts else 0)
        for h in self.search_hits:
            # the flag is about OUR pages: a share button points at facebook.com,
            # and whether anything links to Facebook is none of our business
            if h.kind == "mention" or not self.in_scope(h.absolute):
                h.no_internal = False
                continue
            h.no_internal = visible_per_target.get(url_key(h.absolute), 0) == 0

        # Inbound counts, reported the way established crawlers do it: pages
        # that genuinely link here, and how many of those actually pass weight.
        direct = [h for h in self.search_hits if h.kind == "direct"]
        self.inbound_pages = len({url_key(h.page) for h in direct})
        # a nofollow link is a link a visitor can click but that passes no
        # weight, so it does not count towards the SEO figure
        self.inbound_pages_canonical = len(
            {url_key(h.page) for h in direct if h.source_canonical and not h.nofollow}
        )
        nofollowed = sum(1 for h in direct if h.nofollow)
        if nofollowed:
            self.log(f"  {nofollowed} of them {pl(nofollowed, 'is', 'are')} rel=nofollow - "
                     f"clickable, but passing no weight")
        mentions = sum(1 for h in self.search_hits if h.kind == "mention")
        selfies = sum(1 for h in self.search_hits if h.kind == "self")

        self.log(f"  {self.inbound_pages} {pl(self.inbound_pages, 'page')} link here "
                 f"({self.inbound_pages_canonical} of them canonical - the number an "
                 f"SEO crawler reports)")
        if mentions:
            self.log(f"  {mentions} {pl(mentions, 'match', 'matches')} only mention the "
                     f"address inside another link (share buttons, redirects) - "
                     f"not a link to it")
        if selfies:
            self.log(f"  {selfies} {pl(selfies, 'match', 'matches')} "
                     f"{pl(selfies, 'is', 'are')} the page referring to itself")
        non_canon = self.inbound_pages - self.inbound_pages_canonical
        if non_canon:
            self.log(f"  {non_canon} source {pl(non_canon, 'page')} "
                     f"{pl(non_canon, 'is', 'are')} non-canonical (usually pagination) - "
                     f"search engines discount {pl(non_canon, 'it', 'them')}")

        orphaned = {url_key(h.absolute) for h in self.search_hits if h.no_internal}
        if orphaned:
            self.log(f"  NOTE: {len(orphaned)} of the matched "
                     f"{pl(len(orphaned), 'address', 'addresses')} "
                     f"{pl(len(orphaned), 'has', 'have')} no internal links - only "
                     f"technical tags (<head>, scripts) point there, so nothing on "
                     f"the site actually links to it")

        # Check whether the found link itself is alive: people often search for
        # an address that no longer exists.
        unique = list({h.absolute for h in self.search_hits})
        if len(unique) > 500:
            self.log(f"  {len(unique)} distinct addresses found - skipping status checks "
                     f"(too many), use the 'Broken links' mode instead")
            return
        self.log(f"Checking whether the found addresses work ({len(unique)}) ...")
        with ThreadPoolExecutor(max_workers=self.opts.workers) as pool:
            for url, status in zip(unique, pool.map(self.check_link, unique)):
                self.status_cache[url] = status
        for h in self.search_hits:
            h.status = self.status_cache.get(h.absolute)

        dead = {
            u for u in unique
            if isinstance(self.status_cache.get(u), str)
            or (isinstance(self.status_cache.get(u), int) and self.status_cache[u] >= 400)
        }
        if dead:
            self.log(f"  NOTE: {len(dead)} of the found addresses "
                     f"{pl(len(dead), 'does', 'do')} not work:")
            for u in list(dead)[:5]:
                self.log(f"    {u} - {describe_status(self.status_cache[u])}")

    # -- phase 3: broken links -----------------------------------------------

    def check_link(self, url: str, attempt: int = 0) -> object:
        try:
            r = self.session.head(url, timeout=TIMEOUT, allow_redirects=True)
            if r.status_code in (400, 403, 405, 501) or r.status_code >= 500:
                # some servers do not handle HEAD - retry with GET
                r = self.session.get(url, timeout=TIMEOUT, allow_redirects=True, stream=True)
                r.close()
            if r.status_code in RETRY_STATUSES and attempt < RETRIES and not self.stopped():
                time.sleep(RETRY_PAUSE)
                return self.check_link(url, attempt + 1)
            return r.status_code
        except Exception as e:
            import requests.exceptions as rex
            if isinstance(e, (rex.Timeout, rex.ConnectionError)) \
                    and attempt < RETRIES and not self.stopped():
                time.sleep(RETRY_PAUSE)
                return self.check_link(url, attempt + 1)
            return describe_exception(e)

    def run_broken_check(self) -> None:
        targets: dict[str, list[LinkHit]] = {}
        for lk in self.all_links:
            if lk.tag == RAW_TAG:
                continue
            # page links (<a> and <head> pointers such as hreflang/canonical) are
            # always checked; images, css and scripts only on request
            if not self.opts.check_assets and not points_at_page(lk.tag, lk.rel):
                continue
            internal = self.in_scope(lk.absolute)
            if not internal and not self.opts.check_external:
                continue
            targets.setdefault(lk.absolute, []).append(lk)

        # anything already downloaded during the crawl is not requested again
        to_check = [u for u in targets if u not in self.status_cache]
        self.log(f"Checking {len(to_check)} unique links "
                 f"({len(targets) - len(to_check)} already known from the crawl) ...")

        done = 0
        with ThreadPoolExecutor(max_workers=self.opts.workers) as pool:
            for url, status in zip(to_check, pool.map(self.check_link, to_check)):
                self.status_cache[url] = status
                done += 1
                if done % 25 == 0 or done == len(to_check):
                    self.progress("check", done, len(to_check))
                    self.log(f"  checked {done}/{len(to_check)}")
                if self.stopped():
                    break

        for url, hits in targets.items():
            status = self.status_cache.get(url)
            if status is None:
                continue
            is_broken = isinstance(status, str) or (isinstance(status, int) and status >= 400)
            if not is_broken:
                continue
            reason = describe_status(status)
            scope = "internal" if self.in_scope(url) else "external"
            # the same link often sits on a page several times (menu + footer) -
            # collapse those into one row with a counter
            seen_rows: dict[tuple, dict] = {}
            for lk in hits:
                sig = (lk.page, lk.text, lk.tag, lk.rel)
                if sig in seen_rows:
                    seen_rows[sig]["count"] += 1
                    continue
                seen_rows[sig] = {
                    "page": lk.page,
                    "link": url,
                    "status": status if isinstance(status, int) else "ERROR",
                    "reason": reason,
                    "text": lk.text,
                    "tag": lk.tag,
                    "where": lk.where,
                    "visible": lk.visible,
                    "scope": scope,
                    "href": lk.href,
                    "count": 1,
                }
            self.broken.extend(seen_rows.values())

        self.broken.sort(key=lambda r: (str(r["status"]), r["link"]))
        self.log(f"Broken links: {len(self.broken)} "
                 f"({len({b['link'] for b in self.broken})} unique addresses)")

        hidden = [b for b in self.broken if not b["visible"]]
        if hidden:
            uniq = len({b["link"] for b in hidden})
            self.log(f"  of those, {len(hidden)} sit in technical tags "
                     f"(<head> hreflang/canonical, scripts) - {uniq} unique "
                     f"{pl(uniq, 'address', 'addresses')} you cannot spot on the page")

    # -- indexability --------------------------------------------------------

    def load_robots(self) -> None:
        """Reads robots.txt once so we can tell which URLs are disallowed."""
        if self._robots is not None:
            return
        text = self._robots_text
        if text is None:
            try:
                r = self.session.get(urljoin(self.root + "/", "robots.txt"), timeout=TIMEOUT)
                text = r.text if r.status_code == 200 else ""
            except Exception:
                text = ""             # unreachable robots.txt blocks nothing
            self._robots_text = text
        self._robots = RobotsRules(text)

    def judge_indexability(self, page: PageInfo) -> None:
        """
        Decides whether a search engine may index the page, and says why not.

        Checked in the order an engine applies them: a page blocked in
        robots.txt is never fetched, so its meta tags are irrelevant, and a
        canonical pointing elsewhere means the page itself stays out of the
        index even though it answers fine.
        """
        if page.error or not isinstance(page.status, int):
            page.indexable, page.index_status = False, "Error"
            page.index_reason = page.error or "the page could not be fetched"
            return

        if page.status != 200:
            page.indexable, page.index_status = False, f"HTTP {page.status}"
            page.index_reason = describe_status(page.status)
            return

        if "html" not in (page.content_type or "").lower():
            page.indexable, page.index_status = True, "Not a page"
            page.index_reason = (f"a file rather than a page — served as "
                                 f"{page.content_type or 'an unknown type'}")
            return

        if self._robots and not self._robots.allowed(page.url):
            page.indexable, page.index_status = False, "Blocked by robots.txt"
            page.index_reason = ("robots.txt forbids crawlers from opening this "
                                 "address, so it never reaches the index")
            return

        # "noindex" and the shorthand "none" both keep a page out of the index.
        # Name the tag that actually carries it - sending someone to hunt for a
        # meta tag when the directive came in a header wastes their time.
        def blocks_indexing(value: str) -> bool:
            value = (value or "").lower()
            return "noindex" in value or bool(re.search(r"\bnone\b", value))

        if blocks_indexing(page.x_robots) or blocks_indexing(page.meta_robots):
            if blocks_indexing(page.x_robots):
                # a header is invisible in the page source, so say where to look
                source = "sent by the server as an HTTP header X-Robots-Tag"
                value = (page.x_robots or "").strip()
            else:
                source = 'in the page code: <meta name="robots">'
                value = (page.meta_robots or "").strip()

            # "none" is shorthand for "noindex, nofollow" and reads like the
            # exact opposite to anyone who has not met it before
            shorthand = (" — \"none\" is shorthand for \"noindex, nofollow\""
                         if re.search(r"\bnone\b", value.lower())
                         and "noindex" not in value.lower() else "")

            page.indexable, page.index_status = False, "Noindex"
            page.index_reason = (
                f"the page asks search engines to skip it: {value}{shorthand} "
                f"({source})"
            )
            if "hs_preview=" in page.url.lower():
                page.index_reason += (
                    " — this is a preview link of an unpublished draft, "
                    "those are never indexed"
                )
            return

        if page.canonical and url_key(page.canonical) != url_key(page.url):
            page.indexable, page.index_status = False, "Canonicalised"
            page.index_reason = (
                f"the page declares another address as the original, so search "
                f"engines index that one instead: {page.canonical}"
            )
            return

        page.indexable, page.index_status, page.index_reason = True, "Indexable", ""

    def run_indexability_check(self) -> None:
        """Classifies every crawled page. Costs nothing extra - no requests."""
        self.load_robots()
        for page in self.pages.values():
            self.judge_indexability(page)

        # How many pages link TO each page. Not to be confused with the number
        # of links ON it - the two answer opposite questions.
        inbound = self.count_inbound_links()
        for page in self.pages.values():
            page.inbound = len(inbound.get(url_key(page.url), ()))

        blocked = [p for p in self.pages.values() if not p.indexable]
        self.log(f"Indexability: {len(self.pages) - len(blocked)} of {len(self.pages)} "
                 f"pages can be indexed, {len(blocked)} cannot")

        by_reason: dict[str, int] = {}
        for p in blocked:
            by_reason[p.index_status] = by_reason.get(p.index_status, 0) + 1
        for status, count in sorted(by_reason.items(), key=lambda kv: -kv[1]):
            self.log(f"    {count} x {status}")

        # a page that is both listed in the sitemap and non-indexable is a
        # contradiction worth pointing at: the site tells Google to index it
        # and not to index it at the same time
        contradictory = [
            p for p in blocked
            if url_key(p.url) in self.sitemap_keys and p.index_status in ("Noindex", "Canonicalised")
        ]
        if contradictory:
            self.log(f"  NOTE: {len(contradictory)} of them are listed in sitemap.xml "
                     f"while telling search engines to skip them - the sitemap and the "
                     f"page contradict each other")
            for p in contradictory[:5]:
                self.log(f"    {p.url} - {p.index_status}")

    # -- phase 4: pages nothing links to -------------------------------------

    def count_inbound_links(self) -> dict[str, set[str]]:
        """
        For every address: which pages link to it with a normal link.

        Only <a>/<area>/<form> counts. hreflang and canonical are deliberately
        excluded - search engines do not treat them as internal linking, and a
        page reachable only through them is exactly what we want to surface.
        """
        inbound: dict[str, set[str]] = {}
        for lk in self.all_links:
            if lk.tag not in NAVIGATION_TAGS:
                continue
            if not self.in_scope(lk.absolute):
                continue
            src, dst = url_key(lk.page), url_key(lk.absolute)
            if src == dst:          # a page linking to itself proves nothing
                continue
            inbound.setdefault(dst, set()).add(lk.page)
        return inbound

    def run_orphan_check(self) -> None:
        """Finds pages that exist but that no page on the site links to."""
        if not self.opts.use_sitemap:
            self.log("Orphan check needs sitemap.xml: a page found by crawling is, "
                     "by definition, linked from somewhere. Skipping.")
            return
        if not self.sitemap_count:
            self.log("Orphan check skipped: this site has no sitemap, so every page "
                     "here was found by following links - none of them can be an "
                     "orphan. Nothing to report.")
            return
        if self.opts.limit or self.opts.max_depth is not None:
            self.log("NOTE: crawl was limited, so this list may include pages we "
                     "simply never reached. Run without limits for a reliable answer.")

        inbound = self.count_inbound_links()
        root_key = url_key(self.root)

        skipped_assets = 0
        for page in self.pages.values():
            if page.status != 200:          # broken pages belong to the other report
                continue
            # Sitemaps sometimes list images and PDFs (HubSpot exports /hubfs/*.jpg
            # that way). Those are files, not pages - nobody expects a link to them.
            if "html" not in (page.content_type or "").lower() or not is_page_like(page.url):
                skipped_assets += 1
                continue
            key = url_key(page.url)
            if key == root_key:             # the home page is the entry point
                continue
            sources = inbound.get(key, set())
            if sources:
                continue
            self.orphans.append({
                "url": page.url,
                "title": page.title,
                "links_out": page.links_count,
                "in_sitemap": key in self.sitemap_keys,
            })

        self.orphans.sort(key=lambda r: r["url"])
        if skipped_assets:
            self.log(f"  ignored {skipped_assets} non-HTML entries from the sitemap "
                     f"(images, PDFs and the like)")
        self.log(f"Pages with no internal links: {len(self.orphans)}")
        if self.orphans:
            self.log("  nothing on the site links to these - they are reachable "
                     "only by knowing the address:")
            for o in self.orphans[:5]:
                self.log(f"    {o['url']}")
            if len(self.orphans) > 5:
                self.log(f"    ... and {len(self.orphans) - 5} more")
            self.log("  NOTE: if the site builds its menus or listings with JavaScript, "
                     "those links are invisible to any crawler - spot-check a few "
                     "addresses before treating the whole list as a problem")

    # -- run -----------------------------------------------------------------

    def run(self) -> dict:
        self.started_at = datetime.now()
        mode = self.opts.mode
        self.log(f"=== Start: {self.root} | mode: {mode} ===")

        self.crawl()

        # costs no requests - every signal came with the pages we already fetched
        if not self.stopped():
            self.run_indexability_check()

        if not self.stopped() and mode in ("search", "full") and self.opts.query:
            self.run_search()

        if not self.stopped() and mode in ("broken", "full"):
            self.run_broken_check()
            if self.opts.find_orphans:
                self.run_orphan_check()

        self.finished_at = datetime.now()
        elapsed = (self.finished_at - self.started_at).total_seconds()
        self.log(f"=== Done in {elapsed:.0f}s ===")
        return self.summary()

    def summary(self) -> dict:
        return {
            "domain": self.root,
            "mode": self.opts.mode,
            "query": self.opts.query,
            "pages": len(self.pages),
            "links": len([l for l in self.all_links if l.tag != RAW_TAG]),
            "hits": len(self.search_hits),
            "hit_pages": len({h.page for h in self.search_hits}),
            "inbound_pages": self.inbound_pages,
            "inbound_pages_canonical": self.inbound_pages_canonical,
            "mentions": sum(1 for h in self.search_hits if h.kind == "mention"),
            "broken": len(self.broken),
            "broken_unique": len({b["link"] for b in self.broken}),
            "orphans": len(self.orphans),
            "orphans_checked": self.opts.find_orphans,
            "non_indexable": sum(1 for p in self.pages.values() if not p.indexable),
            "only_non_indexable": self.opts.only_non_indexable,
            "elapsed": (
                (self.finished_at - self.started_at).total_seconds()
                if self.started_at and self.finished_at else 0
            ),
            "stopped": self.stopped(),
        }


# ----------------------------------------------------------------------------
# Excel export
# ----------------------------------------------------------------------------

def export_xlsx(auditor: SiteAuditor, path: str) -> str:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2F4858")

    wb = openpyxl.Workbook()

    def make_sheet(title: str, headers: list[str], rows: Iterable[tuple], widths: list[int], first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title[:31]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center")
        for row in rows:
            ws.append(row)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        return ws

    s = auditor.summary()
    make_sheet(
        "Summary",
        ["Metric", "Value"],
        [
            ("Domain", s["domain"]),
            ("Mode", s["mode"]),
            ("Searched for", s["query"] or "-"),
            ("Pages crawled", s["pages"]),
            ("Links collected", s["links"]),
            ("Matches found", s["hits"]),
            ("Pages with matches", s["hit_pages"]),
            ("Pages that link here", s["inbound_pages"]),
            ("... of them canonical (SEO count)", s["inbound_pages_canonical"]),
            ("Mentions inside other links (not links here)", s["mentions"]),
            ("Broken links (occurrences)", s["broken"]),
            ("Broken links (unique)", s["broken_unique"]),
            ("Pages with no internal links", s["orphans"] if s["orphans_checked"] else "not checked"),
            ("Pages search engines will skip", s["non_indexable"]),
            ("Duration, seconds", round(s["elapsed"])),
            ("Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ],
        [32, 70],
        first=True,
    )

    if auditor.search_hits:
        make_sheet(
            "Where the link was found",
            ["Page holding the link", "Full link", "Match type", "Source page",
             "Source indexable", "Target indexable", "Why target is skipped",
             "Where exactly", "On the page / technical", "Link status",
             "Link or button text", "href as written", "Source context"],
            [
                (h.page, h.absolute, MATCH_KINDS[h.kind],
                 "canonical" if h.source_canonical else "non-canonical (discounted)",
                 h.source_index_status or "not crawled",
                 h.target_index_status or "not crawled",
                 h.target_index_reason,
                 h.where, "on the page" if h.visible else "technical",
                 describe_status(h.status) if h.status is not None else "not checked",
                 h.text, h.href, h.context)
                for h in auditor.search_hits
            ],
            [50, 50, 32, 26, 20, 20, 54, 42, 22, 36, 24, 40, 50],
        )

    if auditor.broken:
        make_sheet(
            "Broken links",
            ["Page holding the link", "Broken link", "Code", "Reason",
             "Where exactly", "On the page / technical", "Link or button text",
             "Scope", "Times on page"],
            [
                (b["page"], b["link"], b["status"], b["reason"], b["where"],
                 "on the page" if b["visible"] else "technical",
                 b["text"], b["scope"], b.get("count", 1))
                for b in auditor.broken
            ],
            [52, 52, 8, 42, 46, 22, 28, 10, 14],
        )

    if auditor.orphans:
        make_sheet(
            "No internal links",
            ["Page URL", "Title", "Links out of this page", "Listed in sitemap"],
            [
                (o["url"], o["title"], o["links_out"], "yes" if o["in_sitemap"] else "no")
                for o in auditor.orphans
            ],
            [70, 55, 22, 18],
        )

    make_sheet(
        "All pages",
        ["Page URL", "Status", "Indexable", "Why not", "Title",
         "Links pointing here (inbound)", "Links placed on this page (outbound)",
         "Content type", "Error"],
        [
            (p.url, p.status, p.index_status, p.index_reason, p.title,
             p.inbound, p.links_count, p.content_type, p.error)
            for p in sorted(auditor.pages.values(), key=lambda x: x.url)
        ],
        [58, 10, 22, 50, 44, 28, 32, 18, 30],
    )

    wb.save(path)
    return path
